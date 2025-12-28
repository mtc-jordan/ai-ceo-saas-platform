"""
Export Service for generating PDF and Excel reports
"""

import io
import json
from datetime import datetime
from typing import Dict, Any, List, Optional
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


class PDFReport(FPDF):
    """Custom PDF class for Lean Six Sigma reports"""
    
    def __init__(self):
        super().__init__()
        self.set_auto_page_break(auto=True, margin=15)
    
    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'AI CEO Platform - Lean Six Sigma Report', 0, 1, 'C')
        self.set_font('Arial', '', 8)
        self.cell(0, 5, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', 0, 1, 'C')
        self.ln(5)
    
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
    
    def chapter_title(self, title: str):
        self.set_font('Arial', 'B', 14)
        self.set_fill_color(41, 128, 185)
        self.set_text_color(255, 255, 255)
        self.cell(0, 10, title, 0, 1, 'L', fill=True)
        self.set_text_color(0, 0, 0)
        self.ln(5)
    
    def section_title(self, title: str):
        self.set_font('Arial', 'B', 11)
        self.set_text_color(41, 128, 185)
        self.cell(0, 8, title, 0, 1, 'L')
        self.set_text_color(0, 0, 0)
    
    def body_text(self, text: str):
        self.set_font('Arial', '', 10)
        self.multi_cell(0, 6, text)
        self.ln(3)
    
    def add_table(self, headers: List[str], data: List[List[str]], col_widths: Optional[List[int]] = None):
        """Add a table to the PDF"""
        self.set_font('Arial', 'B', 9)
        self.set_fill_color(240, 240, 240)
        
        if col_widths is None:
            col_widths = [190 // len(headers)] * len(headers)
        
        # Headers
        for i, header in enumerate(headers):
            self.cell(col_widths[i], 8, header, 1, 0, 'C', fill=True)
        self.ln()
        
        # Data rows
        self.set_font('Arial', '', 9)
        for row in data:
            for i, cell in enumerate(row):
                self.cell(col_widths[i], 7, str(cell)[:30], 1, 0, 'L')
            self.ln()
        self.ln(5)
    
    def add_metric_box(self, label: str, value: str, color: tuple = (46, 204, 113)):
        """Add a metric highlight box"""
        self.set_fill_color(*color)
        self.set_text_color(255, 255, 255)
        self.set_font('Arial', 'B', 12)
        self.cell(60, 15, f'{label}: {value}', 0, 0, 'C', fill=True)
        self.set_text_color(0, 0, 0)


class ExportService:
    """Service for exporting Lean Six Sigma data to PDF and Excel"""
    
    def __init__(self):
        pass
    
    # ==================== PDF Exports ====================
    
    def export_dmaic_project_pdf(self, project: Dict[str, Any]) -> bytes:
        """Export DMAIC project to PDF"""
        pdf = PDFReport()
        pdf.add_page()
        
        # Title
        pdf.chapter_title(f"DMAIC Project: {project.get('title', 'Untitled')}")
        
        # Project Overview
        pdf.section_title("Project Overview")
        pdf.body_text(f"Status: {project.get('status', 'N/A')}")
        pdf.body_text(f"Current Phase: {project.get('current_phase', 'N/A').upper()}")
        pdf.body_text(f"Belt Level: {project.get('belt_level', 'N/A')}")
        pdf.body_text(f"Priority: {project.get('priority', 'N/A')}")
        
        # Problem Statement
        pdf.section_title("Problem Statement")
        pdf.body_text(project.get('problem_statement', 'Not defined'))
        
        # Goal
        pdf.section_title("Goal Statement")
        pdf.body_text(project.get('goal', 'Not defined'))
        
        # Scope
        pdf.section_title("Scope")
        pdf.body_text(project.get('scope', 'Not defined'))
        
        # Metrics
        pdf.section_title("Metrics")
        metrics_data = [
            ["Metric", "Baseline", "Target", "Current"],
            ["Primary", 
             str(project.get('baseline_metric', 'N/A')),
             str(project.get('target_metric', 'N/A')),
             str(project.get('current_metric', 'N/A'))]
        ]
        pdf.add_table(metrics_data[0], metrics_data[1:])
        
        # Financial Impact
        pdf.section_title("Financial Impact")
        pdf.body_text(f"Estimated Savings: ${project.get('estimated_savings', 0):,.2f}")
        pdf.body_text(f"Actual Savings: ${project.get('actual_savings', 0):,.2f}")
        
        # Timeline
        pdf.section_title("Timeline")
        pdf.body_text(f"Start Date: {project.get('start_date', 'N/A')}")
        pdf.body_text(f"Target Completion: {project.get('target_completion', 'N/A')}")
        
        return pdf.output(dest='S').encode('latin-1')
    
    def export_capability_analysis_pdf(self, analysis: Dict[str, Any]) -> bytes:
        """Export process capability analysis to PDF"""
        pdf = PDFReport()
        pdf.add_page()
        
        pdf.chapter_title("Process Capability Analysis Report")
        
        # Statistics
        stats = analysis.get('statistics', {})
        pdf.section_title("Descriptive Statistics")
        stats_data = [
            ["Statistic", "Value"],
            ["Sample Size", str(analysis.get('sample_size', 'N/A'))],
            ["Mean", str(stats.get('mean', 'N/A'))],
            ["Std Dev", str(stats.get('std_dev', 'N/A'))],
            ["Min", str(stats.get('min', 'N/A'))],
            ["Max", str(stats.get('max', 'N/A'))]
        ]
        pdf.add_table(stats_data[0], stats_data[1:], [95, 95])
        
        # Specification Limits
        specs = analysis.get('specification_limits', {})
        pdf.section_title("Specification Limits")
        pdf.body_text(f"USL: {specs.get('usl', 'N/A')}")
        pdf.body_text(f"LSL: {specs.get('lsl', 'N/A')}")
        pdf.body_text(f"Target: {specs.get('target', 'N/A')}")
        
        # Capability Indices
        indices = analysis.get('capability_indices', {})
        pdf.section_title("Capability Indices")
        indices_data = [
            ["Index", "Value", "Interpretation"],
            ["Cp", str(indices.get('cp', 'N/A')), "Potential capability"],
            ["Cpk", str(indices.get('cpk', 'N/A')), "Actual capability"],
            ["Cpu", str(indices.get('cpu', 'N/A')), "Upper capability"],
            ["Cpl", str(indices.get('cpl', 'N/A')), "Lower capability"]
        ]
        pdf.add_table(indices_data[0], indices_data[1:], [40, 40, 110])
        
        # Performance
        perf = analysis.get('performance', {})
        pdf.section_title("Performance Metrics")
        pdf.body_text(f"Sigma Level: {perf.get('sigma_level', 'N/A')}")
        pdf.body_text(f"Expected PPM: {perf.get('ppm_total', 'N/A')}")
        pdf.body_text(f"Expected Yield: {perf.get('yield_percent', 'N/A')}%")
        
        # Interpretation
        interp = analysis.get('interpretation', {})
        pdf.section_title("Interpretation")
        pdf.body_text(f"Rating: {interp.get('cpk_rating', 'N/A')}")
        pdf.body_text(f"Centering: {interp.get('centering', 'N/A')}")
        
        # Recommendations
        pdf.section_title("Recommendations")
        for rec in analysis.get('recommendations', []):
            pdf.body_text(f"• {rec}")
        
        return pdf.output(dest='S').encode('latin-1')
    
    def export_waste_report_pdf(self, waste_data: List[Dict[str, Any]], summary: Dict[str, Any]) -> bytes:
        """Export waste tracking report to PDF"""
        pdf = PDFReport()
        pdf.add_page()
        
        pdf.chapter_title("Waste Tracking Report (TIMWOODS)")
        
        # Summary
        pdf.section_title("Summary")
        pdf.body_text(f"Total Waste Items: {len(waste_data)}")
        pdf.body_text(f"Total Cost Impact: ${summary.get('total_cost', 0):,.2f}")
        pdf.body_text(f"Total Time Impact: {summary.get('total_time', 0)} min/day")
        
        # By Category
        pdf.section_title("Waste by Category")
        category_data = [["Category", "Count", "Cost Impact"]]
        for cat, data in summary.get('by_category', {}).items():
            category_data.append([cat, str(data.get('count', 0)), f"${data.get('cost', 0):,.2f}"])
        pdf.add_table(category_data[0], category_data[1:], [80, 40, 70])
        
        # Detailed List
        pdf.add_page()
        pdf.section_title("Waste Items Detail")
        for waste in waste_data[:20]:  # Limit to 20 items
            pdf.set_font('Arial', 'B', 10)
            pdf.cell(0, 6, f"{waste.get('waste_type', 'Unknown')} - {waste.get('status', 'N/A')}", 0, 1)
            pdf.set_font('Arial', '', 9)
            pdf.body_text(f"Description: {waste.get('description', 'N/A')}")
            pdf.body_text(f"Location: {waste.get('location', 'N/A')} | Cost: ${waste.get('cost_impact', 0):,.2f}")
            pdf.ln(3)
        
        return pdf.output(dest='S').encode('latin-1')
    
    def export_oee_report_pdf(self, oee_data: List[Dict[str, Any]], summary: Dict[str, Any]) -> bytes:
        """Export OEE report to PDF"""
        pdf = PDFReport()
        pdf.add_page()
        
        pdf.chapter_title("OEE (Overall Equipment Effectiveness) Report")
        
        # Summary Metrics
        pdf.section_title("Summary Metrics")
        pdf.ln(5)
        
        # Add metric boxes
        avg_oee = summary.get('average_oee', 0)
        pdf.add_metric_box("Avg OEE", f"{avg_oee:.1f}%", 
                         (46, 204, 113) if avg_oee >= 85 else (241, 196, 15) if avg_oee >= 65 else (231, 76, 60))
        pdf.cell(10, 15, '', 0, 0)
        pdf.add_metric_box("Availability", f"{summary.get('average_availability', 0):.1f}%")
        pdf.cell(10, 15, '', 0, 0)
        pdf.add_metric_box("Performance", f"{summary.get('average_performance', 0):.1f}%")
        pdf.ln(25)
        
        # OEE Formula
        pdf.section_title("OEE Formula")
        pdf.body_text("OEE = Availability × Performance × Quality")
        pdf.body_text("World Class OEE: 85% or higher")
        
        # Equipment Details
        pdf.section_title("Equipment OEE Details")
        equip_data = [["Equipment", "OEE", "Availability", "Performance", "Quality"]]
        for record in oee_data[:15]:
            equip_data.append([
                record.get('equipment_name', 'N/A')[:20],
                f"{record.get('oee', 0):.1f}%",
                f"{record.get('availability', 0):.1f}%",
                f"{record.get('performance', 0):.1f}%",
                f"{record.get('quality', 0):.1f}%"
            ])
        pdf.add_table(equip_data[0], equip_data[1:], [50, 30, 35, 40, 35])
        
        # Recommendations
        pdf.section_title("Recommendations")
        if avg_oee < 65:
            pdf.body_text("• URGENT: OEE is significantly below world-class. Focus on major improvement initiatives.")
        elif avg_oee < 85:
            pdf.body_text("• OEE has room for improvement. Analyze losses and implement targeted improvements.")
        else:
            pdf.body_text("• Excellent OEE performance. Focus on maintaining and continuous improvement.")
        
        return pdf.output(dest='S').encode('latin-1')
    
    # ==================== Excel Exports ====================
    
    def export_dmaic_projects_excel(self, projects: List[Dict[str, Any]]) -> bytes:
        """Export DMAIC projects to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DMAIC Projects"
        
        # Header styling
        header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ["Title", "Phase", "Status", "Belt Level", "Priority", 
                  "Baseline", "Target", "Current", "Est. Savings", "Actual Savings",
                  "Start Date", "Target Completion"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row, project in enumerate(projects, 2):
            ws.cell(row=row, column=1, value=project.get('title', ''))
            ws.cell(row=row, column=2, value=project.get('current_phase', '').upper())
            ws.cell(row=row, column=3, value=project.get('status', ''))
            ws.cell(row=row, column=4, value=project.get('belt_level', ''))
            ws.cell(row=row, column=5, value=project.get('priority', ''))
            ws.cell(row=row, column=6, value=project.get('baseline_metric', ''))
            ws.cell(row=row, column=7, value=project.get('target_metric', ''))
            ws.cell(row=row, column=8, value=project.get('current_metric', ''))
            ws.cell(row=row, column=9, value=project.get('estimated_savings', 0))
            ws.cell(row=row, column=10, value=project.get('actual_savings', 0))
            ws.cell(row=row, column=11, value=project.get('start_date', ''))
            ws.cell(row=row, column=12, value=project.get('target_completion', ''))
        
        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_waste_tracking_excel(self, waste_data: List[Dict[str, Any]]) -> bytes:
        """Export waste tracking to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Waste Tracking"
        
        header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ["Type", "Description", "Location", "Frequency", 
                  "Time Impact (min)", "Cost Impact ($)", "Status", 
                  "Root Cause", "Countermeasure"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, waste in enumerate(waste_data, 2):
            ws.cell(row=row, column=1, value=waste.get('waste_type', ''))
            ws.cell(row=row, column=2, value=waste.get('description', ''))
            ws.cell(row=row, column=3, value=waste.get('location', ''))
            ws.cell(row=row, column=4, value=waste.get('frequency', ''))
            ws.cell(row=row, column=5, value=waste.get('time_impact_minutes', 0))
            ws.cell(row=row, column=6, value=waste.get('cost_impact', 0))
            ws.cell(row=row, column=7, value=waste.get('status', ''))
            ws.cell(row=row, column=8, value=waste.get('root_cause', ''))
            ws.cell(row=row, column=9, value=waste.get('countermeasure', ''))
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="Waste Summary by Type")
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Calculate summary
        waste_summary = {}
        for waste in waste_data:
            wtype = waste.get('waste_type', 'Unknown')
            if wtype not in waste_summary:
                waste_summary[wtype] = {'count': 0, 'cost': 0, 'time': 0}
            waste_summary[wtype]['count'] += 1
            waste_summary[wtype]['cost'] += waste.get('cost_impact', 0)
            waste_summary[wtype]['time'] += waste.get('time_impact_minutes', 0)
        
        ws_summary.cell(row=3, column=1, value="Type")
        ws_summary.cell(row=3, column=2, value="Count")
        ws_summary.cell(row=3, column=3, value="Total Cost")
        ws_summary.cell(row=3, column=4, value="Total Time (min)")
        
        for row, (wtype, data) in enumerate(waste_summary.items(), 4):
            ws_summary.cell(row=row, column=1, value=wtype)
            ws_summary.cell(row=row, column=2, value=data['count'])
            ws_summary.cell(row=row, column=3, value=data['cost'])
            ws_summary.cell(row=row, column=4, value=data['time'])
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_oee_data_excel(self, oee_data: List[Dict[str, Any]]) -> bytes:
        """Export OEE data to Excel with charts"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OEE Data"
        
        header_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ["Date", "Equipment", "Shift", "Availability %", "Performance %", 
                  "Quality %", "OEE %", "Good Units", "Defects"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, record in enumerate(oee_data, 2):
            ws.cell(row=row, column=1, value=record.get('record_date', ''))
            ws.cell(row=row, column=2, value=record.get('equipment_name', ''))
            ws.cell(row=row, column=3, value=record.get('shift', ''))
            ws.cell(row=row, column=4, value=record.get('availability', 0))
            ws.cell(row=row, column=5, value=record.get('performance', 0))
            ws.cell(row=row, column=6, value=record.get('quality', 0))
            ws.cell(row=row, column=7, value=record.get('oee', 0))
            ws.cell(row=row, column=8, value=record.get('good_pieces', 0))
            ws.cell(row=row, column=9, value=record.get('total_pieces', 0) - record.get('good_pieces', 0))
        
        # Add chart if enough data
        if len(oee_data) >= 3:
            chart = LineChart()
            chart.title = "OEE Trend"
            chart.y_axis.title = "OEE %"
            chart.x_axis.title = "Record"
            
            data = Reference(ws, min_col=7, min_row=1, max_row=len(oee_data) + 1)
            chart.add_data(data, titles_from_data=True)
            
            ws.add_chart(chart, "K2")
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_capability_analysis_excel(self, analysis: Dict[str, Any], raw_data: List[float]) -> bytes:
        """Export capability analysis to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Capability Analysis"
        
        # Title
        ws.cell(row=1, column=1, value="Process Capability Analysis")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Statistics
        ws.cell(row=3, column=1, value="Descriptive Statistics")
        ws.cell(row=3, column=1).font = Font(bold=True)
        
        stats = analysis.get('statistics', {})
        ws.cell(row=4, column=1, value="Sample Size")
        ws.cell(row=4, column=2, value=analysis.get('sample_size', 0))
        ws.cell(row=5, column=1, value="Mean")
        ws.cell(row=5, column=2, value=stats.get('mean', 0))
        ws.cell(row=6, column=1, value="Std Dev")
        ws.cell(row=6, column=2, value=stats.get('std_dev', 0))
        ws.cell(row=7, column=1, value="Min")
        ws.cell(row=7, column=2, value=stats.get('min', 0))
        ws.cell(row=8, column=1, value="Max")
        ws.cell(row=8, column=2, value=stats.get('max', 0))
        
        # Specification Limits
        ws.cell(row=10, column=1, value="Specification Limits")
        ws.cell(row=10, column=1).font = Font(bold=True)
        
        specs = analysis.get('specification_limits', {})
        ws.cell(row=11, column=1, value="USL")
        ws.cell(row=11, column=2, value=specs.get('usl', 0))
        ws.cell(row=12, column=1, value="LSL")
        ws.cell(row=12, column=2, value=specs.get('lsl', 0))
        ws.cell(row=13, column=1, value="Target")
        ws.cell(row=13, column=2, value=specs.get('target', 'N/A'))
        
        # Capability Indices
        ws.cell(row=15, column=1, value="Capability Indices")
        ws.cell(row=15, column=1).font = Font(bold=True)
        
        indices = analysis.get('capability_indices', {})
        ws.cell(row=16, column=1, value="Cp")
        ws.cell(row=16, column=2, value=indices.get('cp', 0))
        ws.cell(row=17, column=1, value="Cpk")
        ws.cell(row=17, column=2, value=indices.get('cpk', 0))
        ws.cell(row=18, column=1, value="Cpu")
        ws.cell(row=18, column=2, value=indices.get('cpu', 0))
        ws.cell(row=19, column=1, value="Cpl")
        ws.cell(row=19, column=2, value=indices.get('cpl', 0))
        
        # Performance
        ws.cell(row=21, column=1, value="Performance")
        ws.cell(row=21, column=1).font = Font(bold=True)
        
        perf = analysis.get('performance', {})
        ws.cell(row=22, column=1, value="Sigma Level")
        ws.cell(row=22, column=2, value=perf.get('sigma_level', 0))
        ws.cell(row=23, column=1, value="PPM (Defects)")
        ws.cell(row=23, column=2, value=perf.get('ppm_total', 0))
        ws.cell(row=24, column=1, value="Yield %")
        ws.cell(row=24, column=2, value=perf.get('yield_percent', 0))
        
        # Raw data sheet
        ws_data = wb.create_sheet("Raw Data")
        ws_data.cell(row=1, column=1, value="Measurement")
        for i, value in enumerate(raw_data, 2):
            ws_data.cell(row=i, column=1, value=value)
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_control_chart_excel(self, chart_data: Dict[str, Any]) -> bytes:
        """Export control chart data to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Control Chart"
        
        xbar = chart_data.get('xbar_chart', {})
        r_chart = chart_data.get('r_chart', {})
        
        # X-bar data
        ws.cell(row=1, column=1, value="X-bar Chart Data")
        ws.cell(row=1, column=1).font = Font(bold=True)
        
        ws.cell(row=2, column=1, value="Subgroup")
        ws.cell(row=2, column=2, value="Mean")
        ws.cell(row=2, column=3, value="UCL")
        ws.cell(row=2, column=4, value="CL")
        ws.cell(row=2, column=5, value="LCL")
        
        for i, value in enumerate(xbar.get('data_points', []), 3):
            ws.cell(row=i, column=1, value=i-2)
            ws.cell(row=i, column=2, value=value)
            ws.cell(row=i, column=3, value=xbar.get('ucl', 0))
            ws.cell(row=i, column=4, value=xbar.get('center_line', 0))
            ws.cell(row=i, column=5, value=xbar.get('lcl', 0))
        
        # R chart data (separate sheet)
        ws_r = wb.create_sheet("R Chart")
        ws_r.cell(row=1, column=1, value="R Chart Data")
        ws_r.cell(row=1, column=1).font = Font(bold=True)
        
        ws_r.cell(row=2, column=1, value="Subgroup")
        ws_r.cell(row=2, column=2, value="Range")
        ws_r.cell(row=2, column=3, value="UCL")
        ws_r.cell(row=2, column=4, value="CL")
        ws_r.cell(row=2, column=5, value="LCL")
        
        for i, value in enumerate(r_chart.get('data_points', []), 3):
            ws_r.cell(row=i, column=1, value=i-2)
            ws_r.cell(row=i, column=2, value=value)
            ws_r.cell(row=i, column=3, value=r_chart.get('ucl', 0))
            ws_r.cell(row=i, column=4, value=r_chart.get('center_line', 0))
            ws_r.cell(row=i, column=5, value=r_chart.get('lcl', 0))
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    def export_pareto_analysis_excel(self, pareto_data: Dict[str, Any]) -> bytes:
        """Export Pareto analysis to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pareto Analysis"
        
        chart_data = pareto_data.get('chart_data', {})
        
        ws.cell(row=1, column=1, value="Pareto Analysis")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        headers = ["Category", "Value", "Percentage", "Cumulative %"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
            ws.cell(row=3, column=col).font = Font(bold=True)
        
        categories = chart_data.get('categories', [])
        values = chart_data.get('values', [])
        percentages = chart_data.get('percentages', [])
        cumulative = chart_data.get('cumulative_percentages', [])
        
        for i in range(len(categories)):
            ws.cell(row=i+4, column=1, value=categories[i])
            ws.cell(row=i+4, column=2, value=values[i])
            ws.cell(row=i+4, column=3, value=percentages[i])
            ws.cell(row=i+4, column=4, value=cumulative[i])
        
        # Vital Few summary
        row = len(categories) + 6
        ws.cell(row=row, column=1, value="Vital Few (80% of impact)")
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        for item in pareto_data.get('vital_few', []):
            row += 1
            ws.cell(row=row, column=1, value=item.get('category', ''))
            ws.cell(row=row, column=2, value=item.get('value', 0))
            ws.cell(row=row, column=3, value=f"{item.get('percentage', 0)}%")
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
