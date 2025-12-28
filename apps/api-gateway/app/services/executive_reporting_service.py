"""
Executive Reporting Service
Automated report generation with PDF/Excel/PowerPoint export and scheduling
"""
import os
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import json
from io import BytesIO

# PDF generation
from fpdf import FPDF

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

# OpenAI for AI-generated insights
from openai import OpenAI


class ExecutiveReportGenerator:
    """Generate executive reports in various formats"""
    
    def __init__(self):
        self.client = OpenAI()
        self.output_dir = "/home/ubuntu/ai-ceo-platform/storage/reports"
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generate_pdf_report(self, report_data: Dict[str, Any], report_type: str = "executive") -> str:
        """Generate PDF executive report"""
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Add first page
        pdf.add_page()
        
        # Header
        pdf.set_font("Arial", "B", 24)
        pdf.set_text_color(31, 41, 55)  # Dark gray
        pdf.cell(0, 15, report_data.get("title", "Executive Report"), ln=True, align="C")
        
        # Subtitle
        pdf.set_font("Arial", "", 12)
        pdf.set_text_color(107, 114, 128)  # Gray
        pdf.cell(0, 10, f"Generated: {datetime.utcnow().strftime('%B %d, %Y')}", ln=True, align="C")
        pdf.cell(0, 10, f"Period: {report_data.get('period', 'Current Month')}", ln=True, align="C")
        
        pdf.ln(10)
        
        # Executive Summary
        pdf.set_font("Arial", "B", 16)
        pdf.set_text_color(31, 41, 55)
        pdf.cell(0, 10, "Executive Summary", ln=True)
        
        pdf.set_font("Arial", "", 11)
        pdf.set_text_color(55, 65, 81)
        summary = report_data.get("executive_summary", "No summary available.")
        pdf.multi_cell(0, 7, summary)
        
        pdf.ln(10)
        
        # Key Metrics
        pdf.set_font("Arial", "B", 16)
        pdf.set_text_color(31, 41, 55)
        pdf.cell(0, 10, "Key Performance Indicators", ln=True)
        
        pdf.set_font("Arial", "", 11)
        metrics = report_data.get("key_metrics", [])
        
        # Create metrics table
        col_width = 45
        row_height = 10
        
        # Table header
        pdf.set_fill_color(243, 244, 246)
        pdf.set_font("Arial", "B", 10)
        pdf.cell(col_width * 2, row_height, "Metric", border=1, fill=True)
        pdf.cell(col_width, row_height, "Current", border=1, fill=True, align="C")
        pdf.cell(col_width, row_height, "Change", border=1, fill=True, align="C")
        pdf.ln()
        
        # Table rows
        pdf.set_font("Arial", "", 10)
        for metric in metrics:
            pdf.cell(col_width * 2, row_height, metric.get("name", ""), border=1)
            pdf.cell(col_width, row_height, str(metric.get("value", "")), border=1, align="C")
            
            change = metric.get("change", 0)
            change_text = f"+{change}%" if change > 0 else f"{change}%"
            if change > 0:
                pdf.set_text_color(16, 185, 129)  # Green
            elif change < 0:
                pdf.set_text_color(239, 68, 68)  # Red
            else:
                pdf.set_text_color(55, 65, 81)
            
            pdf.cell(col_width, row_height, change_text, border=1, align="C")
            pdf.set_text_color(55, 65, 81)
            pdf.ln()
        
        pdf.ln(10)
        
        # Highlights
        if report_data.get("highlights"):
            pdf.set_font("Arial", "B", 16)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(0, 10, "Key Highlights", ln=True)
            
            pdf.set_font("Arial", "", 11)
            for highlight in report_data["highlights"]:
                pdf.set_text_color(16, 185, 129)
                pdf.cell(5, 7, chr(149))  # Bullet
                pdf.set_text_color(55, 65, 81)
                pdf.multi_cell(0, 7, highlight)
        
        pdf.ln(5)
        
        # Concerns/Risks
        if report_data.get("concerns"):
            pdf.set_font("Arial", "B", 16)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(0, 10, "Areas of Concern", ln=True)
            
            pdf.set_font("Arial", "", 11)
            for concern in report_data["concerns"]:
                pdf.set_text_color(239, 68, 68)
                pdf.cell(5, 7, chr(149))
                pdf.set_text_color(55, 65, 81)
                pdf.multi_cell(0, 7, concern)
        
        pdf.ln(5)
        
        # Recommendations
        if report_data.get("recommendations"):
            pdf.add_page()
            pdf.set_font("Arial", "B", 16)
            pdf.set_text_color(31, 41, 55)
            pdf.cell(0, 10, "Recommendations", ln=True)
            
            pdf.set_font("Arial", "", 11)
            for i, rec in enumerate(report_data["recommendations"], 1):
                pdf.set_font("Arial", "B", 11)
                pdf.cell(0, 7, f"{i}. {rec.get('title', '')}", ln=True)
                pdf.set_font("Arial", "", 10)
                pdf.set_text_color(107, 114, 128)
                pdf.multi_cell(0, 6, rec.get("description", ""))
                pdf.set_text_color(55, 65, 81)
                pdf.ln(3)
        
        # Footer on each page
        pdf.set_y(-15)
        pdf.set_font("Arial", "I", 8)
        pdf.set_text_color(156, 163, 175)
        pdf.cell(0, 10, f"AI CEO Platform - Confidential | Page {pdf.page_no()}", align="C")
        
        # Save PDF
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_report_{timestamp}.pdf"
        filepath = os.path.join(self.output_dir, filename)
        pdf.output(filepath)
        
        return filepath
    
    def generate_excel_report(self, report_data: Dict[str, Any], report_type: str = "executive") -> str:
        """Generate Excel report with charts"""
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        subheader_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        
        # Title
        ws_summary['A1'] = report_data.get("title", "Executive Report")
        ws_summary['A1'].font = Font(bold=True, size=18)
        ws_summary['A2'] = f"Generated: {datetime.utcnow().strftime('%B %d, %Y')}"
        ws_summary['A3'] = f"Period: {report_data.get('period', 'Current Month')}"
        
        # Key Metrics
        ws_summary['A5'] = "Key Performance Indicators"
        ws_summary['A5'].font = subheader_font
        
        metrics = report_data.get("key_metrics", [])
        headers = ["Metric", "Current Value", "Previous Value", "Change %", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row, metric in enumerate(metrics, 7):
            ws_summary.cell(row=row, column=1, value=metric.get("name", "")).border = border
            ws_summary.cell(row=row, column=2, value=metric.get("value", "")).border = border
            ws_summary.cell(row=row, column=3, value=metric.get("previous", "")).border = border
            ws_summary.cell(row=row, column=4, value=metric.get("change", 0)).border = border
            ws_summary.cell(row=row, column=5, value=metric.get("status", "")).border = border
        
        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 15
        ws_summary.column_dimensions['D'].width = 12
        ws_summary.column_dimensions['E'].width = 12
        
        # Data Sheet
        ws_data = wb.create_sheet("Detailed Data")
        
        # Add monthly data
        monthly_data = report_data.get("monthly_data", [])
        if monthly_data:
            data_headers = ["Month", "Revenue", "Customers", "MRR", "Churn Rate"]
            for col, header in enumerate(data_headers, 1):
                cell = ws_data.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            for row, data in enumerate(monthly_data, 2):
                ws_data.cell(row=row, column=1, value=data.get("month", ""))
                ws_data.cell(row=row, column=2, value=data.get("revenue", 0))
                ws_data.cell(row=row, column=3, value=data.get("customers", 0))
                ws_data.cell(row=row, column=4, value=data.get("mrr", 0))
                ws_data.cell(row=row, column=5, value=data.get("churn_rate", 0))
            
            # Add chart
            chart = LineChart()
            chart.title = "Revenue Trend"
            chart.y_axis.title = "Revenue ($)"
            chart.x_axis.title = "Month"
            
            data_ref = Reference(ws_data, min_col=2, min_row=1, max_col=2, max_row=len(monthly_data) + 1)
            cats_ref = Reference(ws_data, min_col=1, min_row=2, max_row=len(monthly_data) + 1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            ws_data.add_chart(chart, "G2")
        
        # Save Excel
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type}_report_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    async def generate_ai_insights(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate AI-powered insights for the report"""
        
        prompt = f"""Analyze this business data and provide executive insights:

Data:
{json.dumps(data, indent=2)}

Provide:
1. Executive Summary (2-3 sentences)
2. Key Highlights (3-5 positive points)
3. Areas of Concern (2-3 risks or issues)
4. Strategic Recommendations (3-4 actionable items)
5. Outlook (brief future perspective)

Respond in JSON format:
{{
    "executive_summary": "...",
    "highlights": ["...", "..."],
    "concerns": ["...", "..."],
    "recommendations": [
        {{"title": "...", "description": "...", "priority": "high/medium/low"}}
    ],
    "outlook": "..."
}}"""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[
                    {"role": "system", "content": "You are a business analyst providing executive-level insights and recommendations."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                response_format={"type": "json_object"}
            )
            
            return json.loads(response.choices[0].message.content)
            
        except Exception as e:
            return {
                "executive_summary": "Unable to generate AI insights.",
                "highlights": [],
                "concerns": [],
                "recommendations": [],
                "outlook": "",
                "error": str(e)
            }


class ReportScheduler:
    """Schedule and manage automated report generation"""
    
    def __init__(self):
        self.schedules = []
    
    def create_schedule(self, schedule_config: Dict[str, Any]) -> Dict[str, Any]:
        """Create a new report schedule"""
        schedule_id = len(self.schedules) + 1
        
        schedule = {
            "id": schedule_id,
            "name": schedule_config.get("name", f"Report Schedule {schedule_id}"),
            "report_type": schedule_config.get("report_type", "executive"),
            "frequency": schedule_config.get("frequency", "weekly"),  # daily, weekly, monthly
            "day_of_week": schedule_config.get("day_of_week", 1),  # 1=Monday for weekly
            "day_of_month": schedule_config.get("day_of_month", 1),  # For monthly
            "time": schedule_config.get("time", "09:00"),
            "timezone": schedule_config.get("timezone", "UTC"),
            "recipients": schedule_config.get("recipients", []),
            "format": schedule_config.get("format", ["pdf"]),  # pdf, excel, both
            "include_ai_insights": schedule_config.get("include_ai_insights", True),
            "is_active": True,
            "created_at": datetime.utcnow().isoformat(),
            "last_run": None,
            "next_run": self._calculate_next_run(schedule_config)
        }
        
        self.schedules.append(schedule)
        return schedule
    
    def _calculate_next_run(self, config: Dict[str, Any]) -> str:
        """Calculate next scheduled run time"""
        now = datetime.utcnow()
        frequency = config.get("frequency", "weekly")
        time_str = config.get("time", "09:00")
        hour, minute = map(int, time_str.split(":"))
        
        if frequency == "daily":
            next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(days=1)
        
        elif frequency == "weekly":
            day_of_week = config.get("day_of_week", 1)
            days_ahead = day_of_week - now.weekday()
            if days_ahead <= 0:
                days_ahead += 7
            next_run = now + timedelta(days=days_ahead)
            next_run = next_run.replace(hour=hour, minute=minute, second=0, microsecond=0)
        
        elif frequency == "monthly":
            day_of_month = config.get("day_of_month", 1)
            next_run = now.replace(day=day_of_month, hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                if now.month == 12:
                    next_run = next_run.replace(year=now.year + 1, month=1)
                else:
                    next_run = next_run.replace(month=now.month + 1)
        
        else:
            next_run = now + timedelta(days=1)
        
        return next_run.isoformat()
    
    def get_schedules(self) -> List[Dict[str, Any]]:
        """Get all report schedules"""
        return self.schedules
    
    def update_schedule(self, schedule_id: int, updates: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Update an existing schedule"""
        for schedule in self.schedules:
            if schedule["id"] == schedule_id:
                schedule.update(updates)
                schedule["next_run"] = self._calculate_next_run(schedule)
                return schedule
        return None
    
    def delete_schedule(self, schedule_id: int) -> bool:
        """Delete a schedule"""
        for i, schedule in enumerate(self.schedules):
            if schedule["id"] == schedule_id:
                self.schedules.pop(i)
                return True
        return False


# Mock data for reports
def get_mock_report_data():
    """Generate mock data for executive reports"""
    return {
        "title": "Monthly Executive Report",
        "period": "December 2024",
        "executive_summary": "December showed strong performance with revenue exceeding targets by 12%. Customer acquisition improved while churn remained stable. Key initiatives including the AI Meeting Assistant launch contributed to positive momentum heading into Q1 2025.",
        "key_metrics": [
            {"name": "Monthly Recurring Revenue (MRR)", "value": "$125,000", "previous": "$115,000", "change": 8.7, "status": "On Track"},
            {"name": "Annual Recurring Revenue (ARR)", "value": "$1,500,000", "previous": "$1,380,000", "change": 8.7, "status": "On Track"},
            {"name": "Total Customers", "value": "250", "previous": "235", "change": 6.4, "status": "On Track"},
            {"name": "Net Revenue Retention", "value": "115%", "previous": "112%", "change": 2.7, "status": "Exceeding"},
            {"name": "Customer Acquisition Cost", "value": "$2,500", "previous": "$2,800", "change": -10.7, "status": "Improving"},
            {"name": "Monthly Churn Rate", "value": "1.2%", "previous": "1.5%", "change": -20.0, "status": "Improving"},
            {"name": "NPS Score", "value": "72", "previous": "68", "change": 5.9, "status": "On Track"}
        ],
        "highlights": [
            "Revenue exceeded monthly target by 12%, driven by enterprise deal closures",
            "Successfully launched AI Meeting Assistant module with positive user feedback",
            "Customer acquisition cost reduced by 10.7% through improved marketing efficiency",
            "Net Promoter Score improved to 72, highest in company history",
            "Completed SOC 2 Type II certification ahead of schedule"
        ],
        "concerns": [
            "Two enterprise accounts showing early churn signals - proactive outreach initiated",
            "Engineering team bandwidth constrained for Q1 roadmap delivery",
            "Competitive pressure increasing in mid-market segment"
        ],
        "recommendations": [
            {
                "title": "Accelerate Enterprise Sales",
                "description": "Increase enterprise sales team capacity to capitalize on strong pipeline. Current close rate of 35% suggests opportunity to scale.",
                "priority": "high"
            },
            {
                "title": "Invest in Customer Success",
                "description": "Add two customer success managers to improve retention and expansion in accounts showing growth potential.",
                "priority": "high"
            },
            {
                "title": "Product Differentiation",
                "description": "Fast-track AI features roadmap to maintain competitive advantage in the market.",
                "priority": "medium"
            },
            {
                "title": "Operational Efficiency",
                "description": "Implement automation in onboarding process to reduce time-to-value for new customers.",
                "priority": "medium"
            }
        ],
        "monthly_data": [
            {"month": "Jul 2024", "revenue": 95000, "customers": 200, "mrr": 95000, "churn_rate": 2.1},
            {"month": "Aug 2024", "revenue": 98000, "customers": 210, "mrr": 98000, "churn_rate": 1.8},
            {"month": "Sep 2024", "revenue": 105000, "customers": 218, "mrr": 105000, "churn_rate": 2.5},
            {"month": "Oct 2024", "revenue": 110000, "customers": 228, "mrr": 110000, "churn_rate": 1.5},
            {"month": "Nov 2024", "revenue": 115000, "customers": 235, "mrr": 115000, "churn_rate": 1.9},
            {"month": "Dec 2024", "revenue": 125000, "customers": 250, "mrr": 125000, "churn_rate": 1.2}
        ],
        "outlook": "Q1 2025 outlook is positive with strong pipeline and product momentum. Focus on execution of enterprise strategy and maintaining product velocity will be key to achieving growth targets."
    }


def get_mock_schedules():
    """Get mock report schedules"""
    return [
        {
            "id": 1,
            "name": "Weekly Executive Summary",
            "report_type": "executive",
            "frequency": "weekly",
            "day_of_week": 1,
            "time": "09:00",
            "timezone": "America/New_York",
            "recipients": ["ceo@company.com", "cfo@company.com"],
            "format": ["pdf"],
            "include_ai_insights": True,
            "is_active": True,
            "created_at": "2024-11-01T10:00:00Z",
            "last_run": "2024-12-23T09:00:00Z",
            "next_run": "2024-12-30T09:00:00Z"
        },
        {
            "id": 2,
            "name": "Monthly Board Report",
            "report_type": "board",
            "frequency": "monthly",
            "day_of_month": 1,
            "time": "08:00",
            "timezone": "America/New_York",
            "recipients": ["board@company.com"],
            "format": ["pdf", "excel"],
            "include_ai_insights": True,
            "is_active": True,
            "created_at": "2024-10-15T14:00:00Z",
            "last_run": "2024-12-01T08:00:00Z",
            "next_run": "2025-01-01T08:00:00Z"
        },
        {
            "id": 3,
            "name": "Daily Sales Dashboard",
            "report_type": "sales",
            "frequency": "daily",
            "time": "07:00",
            "timezone": "America/New_York",
            "recipients": ["sales-team@company.com"],
            "format": ["excel"],
            "include_ai_insights": False,
            "is_active": True,
            "created_at": "2024-12-01T09:00:00Z",
            "last_run": "2024-12-24T07:00:00Z",
            "next_run": "2024-12-25T07:00:00Z"
        }
    ]
